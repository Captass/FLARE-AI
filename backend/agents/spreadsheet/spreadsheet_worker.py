"""
FLARE AI - Spreadsheet Worker (Excel/XLSX)
TKT-022

Backend worker to create and manipulate XLSX files using openpyxl.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.utils import get_column_letter

class SpreadsheetWorker:
    """
    An agent worker for programmatically handling XLSX spreadsheets.
    """

    def __init__(self, file_path="output.xlsx"):
        self.file_path = file_path
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
        except FileNotFoundError:
            self.workbook = Workbook()
            self.workbook.active.title = "Sheet1"
            self.save()

    def _get_sheet(self, sheet_name=None):
        """Gets a sheet by name, or the active sheet if no name is provided."""
        if sheet_name:
            return self.workbook[sheet_name]
        return self.workbook.active

    def save(self):
        """Saves the workbook to the file_path."""
        self.workbook.save(self.file_path)

    # --- CRUD Operations ---

    def create_sheet(self, sheet_name):
        """Creates a new sheet."""
        if sheet_name in self.workbook.sheetnames:
            return {"status": "error", "message": f"Sheet '{sheet_name}' already exists."}
        self.workbook.create_sheet(title=sheet_name)
        self.save()
        return {"status": "success", "message": f"Sheet '{sheet_name}' created."}

    def read_cell(self, cell, sheet_name=None):
        """Reads a value from a specific cell."""
        sheet = self._get_sheet(sheet_name)
        value = sheet[cell].value
        return {"status": "success", "cell": cell, "value": value}

    def update_cell(self, cell, value, sheet_name=None):
        """Updates a value in a specific cell."""
        sheet = self._get_sheet(sheet_name)
        sheet[cell] = value
        self.save()
        return {"status": "success", "message": f"Cell {cell} updated."}

    def delete_row(self, row_index, sheet_name=None):
        """Deletes a row."""
        sheet = self._get_sheet(sheet_name)
        sheet.delete_rows(row_index, 1)
        self.save()
        return {"status": "success", "message": f"Row {row_index} deleted."}

    def delete_col(self, col_index, sheet_name=None):
        """Deletes a column."""
        sheet = self._get_sheet(sheet_name)
        sheet.delete_cols(col_index, 1)
        self.save()
        return {"status": "success", "message": f"Column {col_index} deleted."}

    # --- Advanced Operations ---

    def set_formula(self, cell, formula, sheet_name=None):
        """Sets a formula for a cell."""
        # Ensure formula starts with =
        if not formula.startswith('='):
            formula = f"={formula}"
        sheet = self._get_sheet(sheet_name)
        sheet[cell] = formula
        self.save()
        return {"status": "success", "message": f"Formula set for cell {cell}."}

    def set_cell_color(self, cell, color="FFFF00", sheet_name=None):
        """
        Sets the background color of a cell.
        Color is a hex value (e.g., FF0000 for red).
        """
        sheet = self._get_sheet(sheet_name)
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        sheet[cell].fill = fill
        self.save()
        return {"status": "success", "message": f"Color set for cell {cell}."}
        
    def set_font_style(self, cell, bold=False, italic=False, color="000000", sheet_name=None):
        """Sets the font style of a cell."""
        sheet = self._get_sheet(sheet_name)
        font = Font(bold=bold, italic=italic, color=color)
        sheet[cell].font = font
        self.save()
        return {"status": "success", "message": f"Font style set for cell {cell}."}

# Example Usage (for testing or direct execution)
if __name__ == "__main__":
    worker = SpreadsheetWorker("test_workbook.xlsx")
    
    # Create
    worker.create_sheet("MyData")
    
    # Update
    worker.update_cell("A1", "Name", sheet_name="MyData")
    worker.update_cell("B1", "Value", sheet_name="MyData")
    worker.update_cell("A2", "Product A", sheet_name="MyData")
    worker.update_cell("B2", 100, sheet_name="MyData")
    worker.update_cell("A3", "Product B", sheet_name="MyData")
    worker.update_cell("B3", 150, sheet_name="MyData")
    
    # Formula
    worker.set_formula("B4", "=SUM(B2:B3)", sheet_name="MyData")
    
    # Style
    worker.set_cell_color("B4", "00FF00", sheet_name="MyData")
    worker.set_font_style("A1:B1", bold=True, sheet_name="MyData") # This is not implemented, just an idea
    
    print("Spreadsheet 'test_workbook.xlsx' created with sample data.")







