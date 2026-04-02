"""
Worker Tableur — FLARE AI.
Gère : Création, édition et formatage de fichiers Excel (.xlsx).
"""
import asyncio
import logging
import json
import uuid
import io
import base64
import operator
import os
import datetime
from typing import TypedDict, Annotated, Sequence, Literal, Optional

from langchain_core.messages import BaseMessage, HumanMessage, AIMessage, SystemMessage
from langchain_core.tools import tool
from langgraph.graph import StateGraph, END
from langgraph.prebuilt import ToolNode
from langchain_core.runnables import RunnableConfig

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, Reference
from openpyxl.utils import get_column_letter

from core.llm_factory import get_llm
from core.memory import SessionMemory
from core.config import settings

logger = logging.getLogger(__name__)

from core.context import (
    current_user_id as _current_user_id,
    current_session_id as _current_session_id,
    current_request_id as _current_request_id,
    generated_images as _generated_images,
    GLOBAL_IMAGE_REGISTRY as _GLOBAL_IMAGE_REGISTRY,
    current_inline_file as _current_inline_file,
)


def _get_inline_file_payload() -> Optional[dict]:
    payload = _current_inline_file.get()
    return payload if isinstance(payload, dict) else None


def _extract_filename_from_source(file_path_or_url: str, fallback: str) -> str:
    if file_path_or_url.startswith("inline://"):
        payload = _get_inline_file_payload()
        return str(payload.get("name") or fallback) if payload else fallback
    if file_path_or_url.startswith("data:"):
        payload = _get_inline_file_payload()
        return str(payload.get("name") or fallback) if payload else fallback
    if "/" in file_path_or_url:
        return file_path_or_url.split("/")[-1].split("?")[0] or fallback
    return file_path_or_url or fallback


def _qualify_chart_range(range_string: str, sheet_name: str) -> str:
    value = str(range_string or "").strip()
    if not value or "!" in value:
        return value

    escaped_sheet = sheet_name.replace("'", "''")
    return f"'{escaped_sheet}'!{value}"


async def _load_xlsx_bytes(file_path_or_url: str) -> bytes:
    if file_path_or_url.startswith("inline://"):
        payload = _get_inline_file_payload()
        if not payload or not payload.get("content"):
            raise ValueError("Aucun fichier inline n'est disponible dans ce contexte.")
        return base64.b64decode(payload["content"])

    if file_path_or_url.startswith("data:"):
        try:
            _, raw_b64 = file_path_or_url.split(",", 1)
            return base64.b64decode(raw_b64)
        except Exception as exc:
            raise ValueError("Le tableur inline est invalide.") from exc

    if file_path_or_url.startswith("http://") or file_path_or_url.startswith("https://"):
        import httpx
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            response = await client.get(file_path_or_url)
            response.raise_for_status()
            return response.content

    with open(file_path_or_url, "rb") as f:
        return f.read()

@tool
async def generate_excel_document(filename: str, sheets_json: str, config: RunnableConfig) -> str:
    """Crée ou met à jour un fichier Microsoft Excel (.xlsx) avec formatage.

    Args:
        filename: Nom du fichier (doit se terminer par .xlsx). Exemple : "Budget.xlsx"
        sheets_json: Chaîne JSON représentant les feuilles et données.
    """
    configurable = config.get("configurable", {}) if config else {}     
    user_id = configurable.get("user_id") or _current_user_id.get() or "anonymous"
    session_id = configurable.get("session_id") or _current_session_id.get() or "default"
    req_id = configurable.get("request_id") or _current_request_id.get()

    if req_id and req_id in _GLOBAL_IMAGE_REGISTRY:
        current_files = _GLOBAL_IMAGE_REGISTRY[req_id]
    else:
        current_files = _generated_images.get() or []

    try:
        sheets_data = json.loads(sheets_json)
    except json.JSONDecodeError as e:
        return f"Erreur : Le format JSON fourni est invalide. Détail : {str(e)}"

    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    wb = Workbook()
    if sheets_data:
        wb.remove(wb.active)

    for sheet_data in sheets_data:
        sheet_name = str(sheet_data.get("name", "Sheet"))[:31]
        ws = wb.create_sheet(title=sheet_name)

        rows = sheet_data.get("rows", [])
        for r_idx, row_data in enumerate(rows, start=1):
            for c_idx, cell_data in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)

                if isinstance(cell_data, dict):
                    val = cell_data.get("value")
                    cell.value = val

                    is_bold = cell_data.get("bold", False)
                    is_italic = cell_data.get("italic", False)
                    # Support font_color et text_color (alias)
                    text_color = cell_data.get("font_color") or cell_data.get("text_color")
                    bg_color = cell_data.get("bg_color")
                    font_size = cell_data.get("font_size", 11)

                    font_kwargs = {'bold': is_bold, 'italic': is_italic, 'size': font_size}
                    if text_color:
                        font_kwargs['color'] = text_color.lstrip('#')
                    cell.font = Font(**font_kwargs)

                    if bg_color:
                        cell.fill = PatternFill(start_color=bg_color.lstrip('#'), end_color=bg_color.lstrip('#'), fill_type="solid")

                    align = cell_data.get("align")
                    wrap = cell_data.get("wrap_text", False)
                    if align or wrap:
                        cell.alignment = Alignment(horizontal=align or "general", wrap_text=wrap)

                    num_fmt = cell_data.get("num_format")
                    if num_fmt:
                        cell.number_format = num_fmt

                    if cell_data.get("border"):
                        thin = Side(style='thin')
                        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                else:
                    cell.value = cell_data

        # Largeurs de colonnes
        col_widths = sheet_data.get("col_widths", [])
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Figer les volets
        freeze = sheet_data.get("freeze_panes")
        if freeze:
            ws.freeze_panes = freeze

        charts = sheet_data.get("charts", [])
        for chart_data in charts:
            chart_type = chart_data.get("type", "bar").lower()
            title = chart_data.get("title", "Graphique")
            data_range_str = chart_data.get("data_range")
            cat_range_str = chart_data.get("category_range")
            anchor = chart_data.get("anchor", "E1")
            chart_width = chart_data.get("width", 400)
            chart_height = chart_data.get("height", 250)

            if not data_range_str or not cat_range_str:
                continue

            try:
                qualified_data_range = _qualify_chart_range(data_range_str, sheet_name)
                qualified_cat_range = _qualify_chart_range(cat_range_str, sheet_name)

                if chart_type == "bar":
                    chart = BarChart()
                elif chart_type == "line":
                    chart = LineChart()
                elif chart_type == "pie":
                    chart = PieChart()
                elif chart_type == "area":
                    chart = AreaChart()
                else:
                    chart = BarChart()

                data = Reference(ws, range_string=qualified_data_range)
                cats = Reference(ws, range_string=qualified_cat_range)

                chart.title = title
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.width = chart_width / 37.8  # pixels -> cm
                chart.height = chart_height / 37.8

                ws.add_chart(chart, anchor)
            except Exception as chart_error:
                logger.warning(
                    "[generate_excel_document] Graphique ignore sur '%s' (%s / %s): %s",
                    sheet_name,
                    data_range_str,
                    cat_range_str,
                    chart_error,
                )
                continue


    doc_io = io.BytesIO()
    wb.save(doc_io)
    doc_bytes = doc_io.getvalue()

    try:
        file_uuid = str(uuid.uuid4())[:8]
        safe_filename = filename.replace(" ", "_")
        storage_path = f"users/{user_id}/conversations/{session_id}/sheet_{file_uuid}_{safe_filename}"

        from core.firebase_client import firebase_storage as storage    
        public_url = storage.upload_file(
            bucket_name=settings.NEXT_PUBLIC_FIREBASE_STORAGE_BUCKET,   
            path=storage_path,
            file_bytes=doc_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if public_url:
            try:
                memory = SessionMemory(session_id=session_id, user_id=user_id)
                memory.save_file_record(file_name=safe_filename, file_url=public_url,
                                        file_type="spreadsheet", mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                                        file_size=len(doc_bytes))       
            except Exception as e:
                logger.error(f"Erreur SQL spreadsheet_worker: {e}")     

        import base64
        b64_doc = base64.b64encode(doc_bytes).decode('utf-8')
        doc_obj = {
            "name": safe_filename,
            "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "url": public_url,
            "data": b64_doc if not public_url else None
        }
        current_files.append(doc_obj)
        if req_id:
            _GLOBAL_IMAGE_REGISTRY[req_id] = current_files
        _generated_images.set(current_files)

        # ── Log coût tableur dans UsageLedger ─────────────────────────
        try:
            from core.database import SessionLocal as _SL, UsageLedger as _UL
            from core.config import MEDIA_PRICING
            import uuid as _uuid
            _db = _SL()
            _db.add(_UL(
                id=str(_uuid.uuid4()),
                user_id=user_id,
                model_name="spreadsheet-worker",
                action_kind="sheet_gen",
                prompt_tokens=0, candidate_tokens=0, total_tokens=0,
                cost_usd=MEDIA_PRICING.get("sheet_gen", 0.005),
                usage_metadata={"filename": safe_filename, "sheets": len(sheets_data)},
            ))
            _db.commit()
            _db.close()
        except Exception as _e:
            logger.warning(f"[generate_excel_document] Erreur log UsageLedger: {_e}")

        if public_url:
            result_message = f"Tableur '{safe_filename}' généré avec succès. URL: {public_url}"
        else:
            result_message = f"Tableur '{safe_filename}' généré. Il a été ajouté aux fichiers de la session."

        return result_message

    except Exception as e:
        logger.error(f"Erreur persistence generate_excel_document: {e}")
        return f"Le document a été généré mais une erreur est survenue lors de l'enregistrement: {e}"

@tool
async def read_excel_document_as_json(file_path_or_url: str) -> str:    
    """Lit un document Excel (.xlsx) et retourne sa structure en JSON, prêt à être modifié et utilisé par generate_excel_document.

    Args:
        file_path_or_url: URL ou chemin local vers le fichier .xlsx
    """
    try:
        wb = load_workbook(io.BytesIO(await _load_xlsx_bytes(file_path_or_url)), data_only=False)
        sheets_json = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = {"name": sheet_name, "rows": []}
            for row in ws.iter_rows():
                row_data = []
                if not any(c.value for c in row):
                    continue
                for cell in row:
                    cell_value = cell.value

                    if isinstance(cell_value, datetime.datetime):       
                        cell_value = cell_value.isoformat()

                    if isinstance(cell_value, str) and cell_value.startswith('='):
                        pass

                    cell_json = {"value": cell_value}
                    if cell.font:
                        if cell.font.bold: cell_json["bold"] = True     
                        if cell.font.italic: cell_json["italic"] = True 
                        if cell.font.color and getattr(cell.font.color, 'rgb', None):
                            rgb = cell.font.color.rgb
                            if isinstance(rgb, str) and len(rgb) in {6, 8} and all(ch in "0123456789ABCDEFabcdef" for ch in rgb):
                                cell_json["text_color"] = rgb[2:] if len(rgb) == 8 else rgb

                    if cell.fill and cell.fill.fill_type == "solid" and cell.fill.fgColor and getattr(cell.fill.fgColor, 'rgb', None):
                        rgb = cell.fill.fgColor.rgb
                        if isinstance(rgb, str) and len(rgb) in {6, 8} and all(ch in "0123456789ABCDEFabcdef" for ch in rgb):
                            cell_json["bg_color"] = rgb[2:] if len(rgb) == 8 else rgb

                    if cell.alignment and cell.alignment.horizontal:    
                        cell_json["align"] = cell.alignment.horizontal  
                    row_data.append(cell_json)
                sheet_data["rows"].append(row_data)

            sheet_data["charts"] = []
            for chart in ws._charts:
                chart_type = "unknown"
                if isinstance(chart, BarChart):
                    chart_type = "bar"
                elif isinstance(chart, LineChart):
                    chart_type = "line"
                elif isinstance(chart, PieChart):
                    chart_type = "pie"

                data_range = ""
                cat_range = ""
                if chart.series:
                    if chart.series[0].values:
                        data_range = chart.series[0].values.string_ref
                    if chart.series[0].labels:
                        cat_range = chart.series[0].labels.string_ref

                sheet_data["charts"].append({
                    "type": chart_type,
                    "title": chart.title or "",
                    "anchor": chart.anchor.upper() if hasattr(chart, 'anchor') else "E1",
                    "data_range": data_range,
                    "category_range": cat_range
                })

            sheets_json.append(sheet_data)

        return json.dumps(sheets_json, indent=2, ensure_ascii=False)    

    except Exception as e:
        logger.error(f"Erreur lors de la lecture JSON du tableur: {e}", exc_info=True)
        return f"Erreur de lecture du tableur en JSON: {e}"


@tool
async def read_excel_document(file_path_or_url: str) -> str:
    """Lit le contenu d'un document Excel (.xlsx) existant pour comprendre sa structure et ses données.

    Args:
        file_path_or_url: URL ou chemin local vers le fichier .xlsx
    """
    try:
        wb = load_workbook(io.BytesIO(await _load_xlsx_bytes(file_path_or_url)), data_only=True)
        result = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result.append(f"--- Feuille: {sheet_name} ---")
            for row in ws.iter_rows(values_only=True):
                if not any(row):
                    continue
                result.append(" | ".join([str(c) if c is not None else "" for c in row]))

        return "\n".join(result)


    except Exception as e:
        return f"Erreur de lecture du tableur Excel: {str(e)}"

@tool
async def update_excel_document(file_path_or_url: str, sheets_json: str, config: RunnableConfig) -> str:
    """Met à jour un document Excel (.xlsx) existant.

    Args:
        file_path_or_url: L'URL ou le chemin du document à modifier.       
        sheets_json: La nouvelle structure JSON du tableur (obtenue après modification du JSON retourné par read_excel_document_as_json).     
    """
    filename = _extract_filename_from_source(file_path_or_url, "Tableur_mis_a_jour.xlsx")
    if filename.startswith("sheet_"):
        parts = filename.split("_", 2)
        if len(parts) >= 3:
            filename = parts[-1]

    result = await generate_excel_document.ainvoke({"filename": filename, "sheets_json": sheets_json}, config=config)

    if file_path_or_url.startswith("http"):
        try:
            from core.firebase_client import firebase_storage
            bucket_url_prefix = f"https://storage.googleapis.com/{settings.NEXT_PUBLIC_FIREBASE_STORAGE_BUCKET}/"
            if file_path_or_url.startswith(bucket_url_prefix):
                path_to_delete = file_path_or_url.replace(bucket_url_prefix, "")
                firebase_storage.delete_file(path_to_delete)
        except Exception as e:
            logger.error(f"Failed to delete old cloud file: {e}")       
            pass

    return f"Mise à jour réussie. {result}"

@tool
async def delete_excel_document(file_path_or_url: str) -> str:
    """Supprime un document Excel (.xlsx) existant.

    Args:
        file_path_or_url: L'URL ou le chemin local du document à supprimer.
    """
    if file_path_or_url.startswith("inline://") or file_path_or_url.startswith("data:"):
        return "Tableur inline supprimé du contexte courant."
    if file_path_or_url.startswith("http"):
        try:
            from core.firebase_client import firebase_storage
            bucket_url_prefix = f"https://storage.googleapis.com/{settings.NEXT_PUBLIC_FIREBASE_STORAGE_BUCKET}/"
            if file_path_or_url.startswith(bucket_url_prefix):
                path_to_delete = file_path_or_url.replace(bucket_url_prefix, "")
                success = firebase_storage.delete_file(path_to_delete)  
                if success:
                    return f"Le document Excel a été supprimé avec succès du stockage cloud."
                else:
                    return "Échec de la suppression du document cloud."
            else:
                return "L'URL fournie ne correspond pas au stockage de l'application. Suppression ignorée."
        except Exception as e:
            logger.error(f"Erreur lors de la suppression cloud: {e}")   
            return f"Erreur lors de la suppression cloud: {e}"
    else:
        try:
            if os.path.exists(file_path_or_url):
                os.remove(file_path_or_url)
                return f"Document local {file_path_or_url} supprimé avec succès."
            else:
                return f"Le fichier local {file_path_or_url} n'existe pas."
        except Exception as e:
            return f"Erreur lors de la suppression locale: {e}"

@tool
async def generate_excel_insights(file_path_or_url: str, data_sheet_name: str, analysis_prompt: str, config: RunnableConfig) -> str:
    """Analyse une feuille de données, génère des insights et crée un nouvel onglet 'Dashboard' avec des graphiques.

    Args:
        file_path_or_url: URL du fichier Excel à analyser.
        data_sheet_name: Nom de la feuille contenant les données.
        analysis_prompt: Le type d'analyse demandé (ex: "Analyse des ventes mensuelles").
    """
    try:
        full_workbook_json_str = await read_excel_document_as_json.ainvoke(file_path_or_url)
        if full_workbook_json_str.startswith("Erreur"):
            return f"Erreur lors de la lecture du fichier pour analyse: {full_workbook_json_str}"
        
        all_sheets = json.loads(full_workbook_json_str)

        data_sheet = next((s for s in all_sheets if s.get("name") == data_sheet_name), None)
        if not data_sheet:
            return f"Erreur: La feuille '{data_sheet_name}' est introuvable dans le document."

        insights_llm = get_llm(temperature=0.4, model_override="gemini-2.5-pro-002")
        
        dashboard_prompt = f"""
        En te basant sur les données suivantes de la feuille '{data_sheet_name}':
        {json.dumps(data_sheet, indent=2)}

        Et la demande d'analyse : "{analysis_prompt}".

        Génère UNIQUEMENT le JSON pour une NOUVELLE feuille de calcul nommée 'Dashboard'.
        Cette feuille doit contenir:
        1. Quelques statistiques clés (KPIs) en haut de la feuille.
        2. Au moins deux graphiques pertinents pour visualiser les données.
        3. Une brève analyse textuelle des tendances observées.

        Le JSON doit suivre ce format:
        {{ "name": "Dashboard", "rows": [ ... ], "charts": [ ... ] }}

        THÈMES DE GRAPHIQUES DISPONIBLES :
        - **Corporate**: en-têtes `1B2A4A` (bleu marine), alternance lignes `F8FAFC`/blanc, accents `2563EB`
        - **Finance**: en-têtes `064E3B` (vert foncé), totaux `FEF3C7` (jaune pâle), négatifs en rouge `DC2626`
        - **Moderne**: en-têtes `0F172A`, accent `F97316` (orange), fond alterné `FFF7ED`/blanc
        """

        raw_response = await insights_llm.ainvoke(dashboard_prompt)
        response_content = raw_response.content

        # Correction TKT-045: Extraire le JSON du contenu du message
        try:
            # Gérer le cas où le LLM enveloppe le JSON dans du markdown
            logger.debug(f"[generate_excel_insights] Raw LLM response: {response_content}")
            json_start = response_content.find('{')
            json_end = response_content.rfind('}') + 1
            if json_start != -1 and json_end != -1:
                clean_json_str = response_content[json_start:json_end]
                dashboard_sheet = json.loads(clean_json_str)
                if not isinstance(dashboard_sheet, dict) or 'name' not in dashboard_sheet or 'rows' not in dashboard_sheet or 'charts' not in dashboard_sheet:
                    raise json.JSONDecodeError("Le JSON du dashboard est invalide.", clean_json_str, 0)
            else:
                raise json.JSONDecodeError("Aucun objet JSON trouvé dans la réponse", response_content, 0)
        except json.JSONDecodeError as e:
            logger.error(f"[generate_excel_insights] Failed to parse JSON response from LLM: {e}")
            return f"Erreur: Le modèle d'analyse n'a pas retourné un JSON de dashboard valide. Réponse reçue: {response_content}"

        all_sheets = [s for s in all_sheets if s.get("name") != "Dashboard"]
        all_sheets.append(dashboard_sheet)
        
        updated_workbook_json = json.dumps(all_sheets, indent=2)

        return await update_excel_document.ainvoke(file_path_or_url=file_path_or_url, sheets_json=updated_workbook_json, config=config)

    except Exception as e:
        logger.error(f"[generate_excel_insights] Erreur: {e}", exc_info=True)
        return f"Une erreur interne est survenue lors de la génération des insights: {e}"

SPREADSHEET_TOOLS = [
    generate_excel_document,
    read_excel_document_as_json,
    read_excel_document,
    update_excel_document,
    delete_excel_document,
    generate_excel_insights,
]

SPREADSHEET_SYSTEM_PROMPT = """Tu es l'Agent Tableur Expert de FLARE AI. Tu produis des fichiers Excel (.xlsx) PROFESSIONNELS, FONCTIONNELS et VISUELLEMENT IMPACTANTS.

━━━ CRUD ━━━
1. **CREATE**: Conçois une structure JSON complète et utilise `generate_excel_document`.
2. **READ**: Utilise `read_excel_document` (aperçu rapide) ou `read_excel_document_as_json` (structure complète).
3. **UPDATE** — Cas le plus important :
   - Si le message contient `[MODIFICATION FICHIER]`, extrais l'URL du fichier Excel depuis ce bloc.
   - Appelle `read_excel_document_as_json` avec cette URL.
   - Si `[SELECTION]Cellules : RxCy[/SELECTION]` est présent : modifie UNIQUEMENT les cellules spécifiées (format RxCy = ligne x, colonne y, base 0). Trouve-les dans le JSON et applique la modification.
   - Sinon : modifie les parties nécessaires.
   - Appelle `update_excel_document` avec l'URL originale + JSON modifié.
4. **DELETE**: Utilise `delete_excel_document`.

━━━ FORMAT JSON ━━━
Le JSON est une liste de feuilles. Structure complète :
```json
[
  {
    "name": "Nom de la feuille",
    "col_widths": [20, 15, 15, 12, 18],
    "freeze_panes": "A2",
    "rows": [
      [
        {"value": "En-tête", "bold": true, "bg_color": "1B2A4A", "font_color": "FFFFFF", "align": "center"},
        {"value": 1250.50, "num_format": "#,##0.00 €", "bold": false},
        {"value": "=SUM(B2:B10)", "bold": true, "bg_color": "F3F4F6"}
      ]
    ],
    "charts": [
      {
        "type": "bar",
        "title": "Ventes mensuelles",
        "anchor": "G2",
        "data_range": "Feuille1!B2:B13",
        "category_range": "Feuille1!A2:A13",
        "width": 400,
        "height": 250
      }
    ]
  }
]
```

━━━ OPTIONS CELLULES ━━━
- `value`: string, number, ou formule Excel (ex: `"=SUM(B2:B10)"`, `"=AVERAGE(C2:C20)"`, `"=IF(D2>100,\"OK\",\"KO\")"`)
- `bold`: true/false
- `italic`: true/false
- `bg_color`: couleur hex sans # (ex: `"1B2A4A"`)
- `font_color`: couleur hex sans #
- `align`: `"left"`, `"center"`, `"right"`
- `num_format`: format Excel (ex: `"#,##0.00 €"`, `"0.0%"`, `"DD/MM/YYYY"`, `"#,##0"`)
- `border`: true (bordure fine autour de la cellule)
- `wrap_text`: true (retour à la ligne automatique)

━━━ OPTIONS GRAPHIQUES ━━━
- `type`: `"bar"`, `"line"`, `"pie"`, `"area"` (scatter non supporté)
- `anchor`: cellule de départ (ex: `"F1"`)
- `width`/`height`: dimensions en pixels (défaut: 400x250)

━━━ EXIGENCES QUALITÉ ━━━
- **En-têtes**: Toujours en gras, fond coloré, texte blanc ou clair. Première ligne de données = en-tête.
- **Formules**: Systématiquement pour totaux, moyennes, pourcentages, calculs dérivés. Utilise des formules réelles.
- **Largeurs colonnes**: Toujours spécifier `col_widths` adapté au contenu (titres longs = 25+, nombres = 12-15).
- **Figer les volets**: `freeze_panes: "A2"` pour fixer l'en-tête quand la feuille a des données.
- **Multi-feuilles**: Créer plusieurs feuilles quand pertinent (ex: "Données" + "Dashboard" + "Résumé").
- **Graphiques**: Toujours inclure au moins 1 graphique pour les tableurs de données (budget, ventes, stats).
- **Données réelles**: Remplis avec des exemples cohérents et réalistes, pas des placeholders "Valeur 1".
- **Nombres formatés**: Utilise `num_format` pour tous les montants, pourcentages, dates.

━━━ THÈMES ━━━
- **Corporate**: en-têtes `1B2A4A` (bleu marine), alternance lignes `F8FAFC`/blanc, accents `2563EB`
- **Finance**: en-têtes `064E3B` (vert foncé), totaux `FEF3C7` (jaune pâle), négatifs en rouge `DC2626`
- **Moderne**: en-têtes `0F172A`, accent `F97316` (orange), fond alterné `FFF7ED`/blanc

━━━ FORMAT DE RÉPONSE ━━━
Réponds brièvement (2 phrases max). Propose 2 suggestions avec `[SUGGESTION: ...]`.
"""

class SpreadsheetWorker:
    """Worker spécialisé dans la création de fichiers Excel."""      

    def __init__(self, model_override: str = None):
        self.tools = SPREADSHEET_TOOLS
        base_llm = get_llm(
            temperature=0.1,
            model_override=model_override or "gemini-2.5-flash",    
        )
        self.llm_forced = base_llm.bind_tools(self.tools, tool_choice="any")
        self.llm = base_llm.bind_tools(self.tools)
        self.tool_node = ToolNode(self.tools)
        self.graph = self._build_graph()
        logger.info(f"[SpreadsheetWorker] Initialisé avec {len(self.tools)} outils")

    def _build_graph(self):
        graph = StateGraph(TypedDict("SpreadsheetState", {"messages": Annotated[Sequence[BaseMessage], operator.add]}))
        graph.add_node("agent", self._call_model)
        graph.add_node("tools", self.tool_node)
        graph.set_entry_point("agent")
        graph.add_conditional_edges("agent", self._should_continue, {"continue": "tools", "end": END})
        graph.add_edge("tools", "agent")
        return graph.compile()

    def _should_continue(self, state) -> Literal["continue", "end"]:    
        last = state["messages"][-1]
        if hasattr(last, "tool_calls") and last.tool_calls:
            return "continue"
        return "end"

    async def _call_model(self, state, config: RunnableConfig = None) -> dict:
        from langchain_core.messages import ToolMessage
        messages = state["messages"]

        if not any(isinstance(m, SystemMessage) for m in messages):     
            messages = [SystemMessage(content=SPREADSHEET_SYSTEM_PROMPT)] + list(messages)

        has_tool_result = any(isinstance(m, ToolMessage) for m in messages)
        llm = self.llm if has_tool_result else self.llm_forced

        for attempt in range(3):
            try:
                return {"messages": [await llm.ainvoke(messages)]}
            except Exception as e:
                err_str = str(e).lower()
                if attempt < 2 and any(k in err_str for k in ["429", "500", "503"]):
                    await asyncio.sleep(2 ** (attempt + 1))
                    continue
                if not has_tool_result and ("tool_choice" in err_str or "any" in err_str or "unsupported" in err_str):
                    logger.warning("[SpreadsheetWorker] tool_choice=any non supporté, injection directive forcée")
                    forced_msgs = list(messages)
                    last = forced_msgs[-1]
                    directive = "\n\nDIRECTIVE OBLIGATOIRE : Tu dois appeler maintenant l'outil tableur approprié. Une réponse textuelle seule n'est pas acceptée."
                    if hasattr(last, "content") and isinstance(last.content, str):
                        forced_msgs[-1] = HumanMessage(content=last.content + directive)
                    return {"messages": [await self.llm.ainvoke(forced_msgs)]}
                raise

    async def run(self, task: str, config: dict = None) -> str:
        prompt_content = f"""[Instructions]
{SPREADSHEET_SYSTEM_PROMPT}
[Fin instructions]

{task}"""








        messages = [HumanMessage(content=prompt_content)]
        result = await self.graph.ainvoke({"messages": messages}, config=config or {})
        last_msg = result["messages"][-1]
        content = getattr(last_msg, "content", "")
        if isinstance(content, list):
            content = " ".join(part.get("text", "") if isinstance(part, dict) else str(part) for part in content)
        return content or "Tâche tableur terminée."






