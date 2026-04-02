from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
OUT_PATH = BASE_DIR / "sheet_template.xlsx"


HEADERS = {
    "contacts": [
        "psid",
        "first_name",
        "last_name",
        "display_name",
        "profile_pic",
        "preferred_language",
        "source_channel",
        "first_contact_at",
        "last_contact_at",
        "current_status",
        "current_mode",
        "assigned_to",
        "tags",
        "contact_details",
        "notes",
    ],
    "conversations": [
        "conversation_id",
        "psid",
        "opened_at",
        "last_message_at",
        "last_customer_message",
        "last_agent_reply",
        "conversation_status",
        "lead_stage",
        "priority",
        "needs_human",
        "current_mode",
        "last_telegram_reason",
        "last_telegram_notified_at",
        "owner",
        "notes",
    ],
    "messages": [
        "message_mid",
        "received_at",
        "psid",
        "direction",
        "customer_name",
        "message_text",
        "reply_text",
        "intent",
        "lead_status",
        "needs_human",
        "order_signal",
        "llm_provider",
        "llm_model",
        "prompt_tokens",
        "output_tokens",
        "total_tokens",
        "estimated_cost_usd",
        "latency_ms",
        "tags",
    ],
    "leads": [
        "lead_id",
        "psid",
        "created_at",
        "customer_name",
        "service_requested",
        "lead_stage",
        "lead_temperature",
        "budget_range",
        "location",
        "contact_details",
        "next_action",
        "next_action_due_at",
        "owner",
        "updated_at",
        "notes",
    ],
    "devis": [
        "quote_id",
        "psid",
        "created_at",
        "customer_name",
        "service_requested",
        "request_summary",
        "quote_status",
        "urgency",
        "responsible_notified_at",
        "contact_details",
        "follow_up_date",
        "amount_estimate",
        "owner",
        "updated_at",
        "notes",
    ],
    "rendez_vous": [
        "meeting_id",
        "psid",
        "created_at",
        "customer_name",
        "request_reason",
        "preferred_date",
        "preferred_time",
        "meeting_status",
        "responsible_notified_at",
        "scheduled_at",
        "contact_details",
        "owner",
        "updated_at",
        "notes",
    ],
    "kpi_journalier": [
        "day_key",
        "new_contacts",
        "active_conversations",
        "messages_received",
        "replies_sent",
        "human_escalations",
        "leads_created",
        "devis_requested",
        "rendez_vous_requested",
        "order_signals",
        "tokens_total",
        "cost_total_usd",
        "avg_latency_ms",
        "notes",
    ],
}


wb = Workbook()
default = wb.active
wb.remove(default)

header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)

for title, headers in HEADERS.items():
    ws = wb.create_sheet(title)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = max(16, len(header) + 2)
    ws.freeze_panes = "A2"

wb.save(OUT_PATH)
print(OUT_PATH)
